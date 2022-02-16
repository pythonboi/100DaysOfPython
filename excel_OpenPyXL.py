# from openpyxl import Workbook

import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\admin\Documents\Test\example.xlsx")

# wb = openpyxl.load_workbook('example.xlsx')

sheet = wb['Sheet3']

print(wb.sheetnames)

type(sheet)

# print(sheet)

print(sheet.title)

anotheSheet = wb.active


print(anotheSheet)